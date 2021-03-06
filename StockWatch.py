import time
import datetime
from yahoo_finance import Share
from openpyxl import Workbook
from openpyxl import load_workbook
from os.path import exists


print 'Welcome to StockWatch'

#excel setup
workbook_path = 'stockprices_.xlsx'

delay_s = 300;
# times work only in Pacific, matched to NASDAQ opening times: http://www.insider-monitor.com/market-hours.php
# 6:30AM to 1PM PST
market_start = datetime.time(06, 30);
# make sure to CHANGE END TIME
market_end = datetime.time(21, 00);
stockWatchList = ['AMZN', 'AAPL']
while True:
    current_time = datetime.datetime.now().time()
    isMarketOpen = (market_start <= current_time <= market_end)
    print 'Is market open right now? ' + isMarketOpen.__str__()
    print 'Current time is: ' + current_time.__str__()
    if isMarketOpen:
        if exists(workbook_path):
            wb = load_workbook(workbook_path)
        else:
            wb = Workbook(workbook_path)
            sheetList = wb.sheetnames
            print sheetList
            print 'Running through items in stock watch list'
            for ticker in stockWatchList:
                if ticker not in sheetList:
                    ws = wb.create_sheet(ticker)
                else:
                    ws = wb.get_sheet_by_name(ticker)
                stockInfo = Share(ticker)
                print stockInfo.get_price()
                ws.append([current_time, stockInfo.get_price()]);
            wb.save(workbook_path)
    print 'resting...' + (delay_s/60).__str__() + ' mins';
    time.sleep(delay_s);
